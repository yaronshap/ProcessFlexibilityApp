# =============================================================================
# ⚠️  WARNING: THIS PROJECT IS CURRENTLY UNDER CONSTRUCTION   ⚠️
# =============================================================================
# This application is in active development. Features may be incomplete,
# unstable, or subject to change. Please use with caution.
# =============================================================================

import streamlit as st
import networkx as nx
import plotly.graph_objects as go
import plotly.express as px
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple
import random
import json
import io
import os
from scipy.optimize import linprog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# CONFIGURATION VARIABLES
# =============================================================================
# Set to True to show Advanced Flexibility Options (2-Flexibility, 3-Flexibility)
# Set to False to hide these options
SHOW_ADVANCED_FLEXIBILITY_OPTIONS = True

# Set page config
st.set_page_config(
    page_title="Process Flexibility Simulator",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Display construction warning banner
st.error("⚠️ **WARNING: This project is currently under construction!** ⚠️\n\n"
         "This application is in active development. Features may be incomplete, "
         "unstable, or subject to change. Please use with caution.", icon="🚧")

# Add custom CSS for smaller connection grid buttons only
st.markdown("""
<style>
    /* Make only connection grid buttons smaller - target by key pattern */
    button[key*="btn_"]:not([data-testid="baseButton-primary"]),
    button[key*="detailed_"]:not([data-testid="baseButton-primary"]) {
        width: 2.5rem !important;
        height: 2rem !important;
        font-size: 0.8rem !important;
        padding: 0.25rem 0.5rem !important;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'connections' not in st.session_state:
    st.session_state.connections = {}
if 'demand_data' not in st.session_state:
    st.session_state.demand_data = {}

def create_bipartite_graph(num_factories: int, num_products: int, connections: Dict[Tuple[int, int], bool]) -> go.Figure:
    """Create an interactive bipartite graph visualization"""
    
    # Calculate positions
    pos = {}
    
    # Position products on the left (sorted with smaller numbers at top)
    for i in range(num_products):
        pos[f"Product_{i}"] = (0, num_products - 1 - i)  # Reverse order: 1 at top
    
    # Position factories on the right (sorted with smaller numbers at top)
    for i in range(num_factories):
        pos[f"Factory_{i}"] = (2, num_factories - 1 - i)  # Reverse order: 1 at top
    
    # Create traces for each edge individually to enable clicking
    traces = []
    
    # Add all possible connections
    for product_idx in range(num_products):
        for factory_idx in range(num_factories):
            is_connected = connections.get((product_idx, factory_idx), False)
            
            # Get positions
            product_pos = pos[f"Product_{product_idx}"]
            factory_pos = pos[f"Factory_{factory_idx}"]
            
            # Create individual edge trace
            edge_trace = go.Scatter(
                x=[product_pos[0], factory_pos[0]], 
                y=[product_pos[1], factory_pos[1]],
                mode='lines',
                line=dict(
                    width=3 if is_connected else 1,
                    color='blue' if is_connected else 'lightgray',
                    dash='solid' if is_connected else 'dash'
                ),
                hoverinfo='text',
                hovertext=f"Product {product_idx+1} ↔ Factory {factory_idx+1}",
                name=f"Edge_{product_idx}_{factory_idx}",
                showlegend=False,
                customdata=[(product_idx, factory_idx)],
                hovertemplate='%{hovertext}<br>Click to toggle<extra></extra>'
            )
            traces.append(edge_trace)
    
    # Create node traces (maintain the sorted order)
    product_x = [pos[f"Product_{i}"][0] for i in range(num_products)]
    product_y = [pos[f"Product_{i}"][1] for i in range(num_products)]
    
    factory_x = [pos[f"Factory_{i}"][0] for i in range(num_factories)]
    factory_y = [pos[f"Factory_{i}"][1] for i in range(num_factories)]
    
    # Product nodes
    product_trace = go.Scatter(
        x=product_x, y=product_y,
        mode='markers+text',
        hoverinfo='text',
        text=[f"Product {i+1}" for i in range(num_products)],
        textposition="middle center",
        marker=dict(size=60, color='lightblue', line=dict(width=3, color='darkblue')),
        name='Products',
        hovertemplate='Product %{text}<extra></extra>',
        showlegend=True
    )
    traces.append(product_trace)
    
    # Factory nodes
    factory_trace = go.Scatter(
        x=factory_x, y=factory_y,
        mode='markers+text',
        hoverinfo='text',
        text=[f"Factory {i+1}" for i in range(num_factories)],
        textposition="middle center",
        marker=dict(size=60, color='lightgreen', line=dict(width=3, color='darkgreen')),
        name='Factories',
        hovertemplate='Factory %{text}<extra></extra>',
        showlegend=True
    )
    traces.append(factory_trace)
    
    # Create figure
    fig = go.Figure(data=traces)
    
    # Update layout
    fig.update_layout(
        title=dict(
            text="Process Flexibility Network",
            font=dict(size=16)
        ),
        showlegend=True,
        hovermode='closest',
        margin=dict(b=20,l=5,r=5,t=40),
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-0.5, 2.5]),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-0.5, max(num_products, num_factories)-0.5]),
        plot_bgcolor='white',
        width=800,
        height=600
    )
    
    return fig

def generate_demand_data(num_products: int, distribution_type: str, params: Dict) -> Dict[int, List[float]]:
    """Generate demand data based on distribution type"""
    demand_data = {}
    
    if distribution_type == "Truncated Normal":
        # Get parameters from the function call
        mean = params.get('mean', 100)
        std = params.get('std', 40)
        min_val = params.get('min', 20)
        max_val = params.get('max', 180)
        
        for i in range(num_products):
            # Generate normal distribution and truncate
            raw_demands = np.random.normal(mean, std, 2000)  # Generate more to account for truncation
            # Truncate to specified range
            truncated_demands = np.clip(raw_demands, min_val, max_val)
            # Round to integers and take first 1000 samples
            demand_data[i] = np.round(truncated_demands[:1000]).astype(int)
    
    return demand_data

def calculate_flexibility_metrics(connections: Dict[Tuple[int, int], bool], 
                                num_products: int, num_factories: int) -> Dict:
    """Calculate flexibility metrics"""
    
    # Count total connections
    total_connections = sum(connections.values())
    max_possible = num_products * num_factories
    
    # Calculate flexibility ratio
    flexibility_ratio = total_connections / max_possible if max_possible > 0 else 0
    
    # Calculate average connections per product
    product_connections = {}
    for (product_idx, factory_idx), connected in connections.items():
        if connected:
            product_connections[product_idx] = product_connections.get(product_idx, 0) + 1
    
    avg_connections_per_product = np.mean(list(product_connections.values())) if product_connections else 0
    
    return {
        'total_connections': total_connections,
        'max_possible': max_possible,
        'flexibility_ratio': flexibility_ratio,
        'avg_connections_per_product': avg_connections_per_product
    }

def solve_maximal_matching(demands: List[float], factory_capacity: float, 
                          connections: Dict[Tuple[int, int], bool], 
                          num_products: int, num_factories: int) -> Tuple[List[float], float, float]:
    """
    Solve maximal matching problem using linear programming
    Returns: (flows, total_shipped, total_lost)
    """
    
    # Create flow variables for each connected edge
    edge_vars = []
    edge_indices = {}
    var_idx = 0
    
    for (product_idx, factory_idx), connected in connections.items():
        if connected:
            edge_indices[(product_idx, factory_idx)] = var_idx
            edge_vars.append(var_idx)
            var_idx += 1
    
    if not edge_vars:
        return [0.0] * len(edge_vars), 0.0, sum(demands)
    
    # Objective: maximize total flow
    c = [-1.0] * len(edge_vars)  # Negative because we minimize
    
    # Constraints
    A_ub = []
    b_ub = []
    
    # Demand constraints: sum of flows from each product <= demand
    for product_idx in range(num_products):
        constraint = [0.0] * len(edge_vars)
        for (p_idx, factory_idx), connected in connections.items():
            if connected and p_idx == product_idx:
                constraint[edge_indices[(p_idx, factory_idx)]] = 1.0
        A_ub.append(constraint)
        b_ub.append(demands[product_idx])
    
    # Capacity constraints: sum of flows to each factory <= capacity
    for factory_idx in range(num_factories):
        constraint = [0.0] * len(edge_vars)
        for (product_idx, p_idx), connected in connections.items():
            if connected and p_idx == factory_idx:
                constraint[edge_indices[(product_idx, p_idx)]] = 1.0
        A_ub.append(constraint)
        b_ub.append(factory_capacity)
    
    # Bounds: flows >= 0
    bounds = [(0, None)] * len(edge_vars)
    
    # Solve
    try:
        result = linprog(c, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method='highs')
        
        if result.success:
            flows = result.x
            total_shipped = sum(flows)
            total_lost = sum(demands) - total_shipped
            return flows, total_shipped, total_lost
        else:
            # If no solution, return zeros
            return [0.0] * len(edge_vars), 0.0, sum(demands)
    except:
        # Fallback: return zeros
        return [0.0] * len(edge_vars), 0.0, sum(demands)

def run_simulation(num_replications: int, num_products: int, num_factories: int, 
                  factory_capacity: float, connections: Dict[Tuple[int, int], bool],
                  demand_params: Dict, progress_bar=None) -> pd.DataFrame:
    """Run the simulation and return results as DataFrame"""
    
    results = []
    
    for rep in range(num_replications):
        # Set seed
        np.random.seed(rep)
        
        # Generate demands using configurable truncated normal
        demands = []
        mean = demand_params.get('mean', 100)
        std = demand_params.get('std', 40)
        min_val = demand_params.get('min', 20)
        max_val = demand_params.get('max', 180)
        
        for _ in range(num_products):
            raw_demand = np.random.normal(mean, std)
            truncated_demand = np.clip(raw_demand, min_val, max_val)
            demands.append(round(truncated_demand))
        
        # Solve maximal matching
        flows, total_shipped, total_lost = solve_maximal_matching(
            demands, factory_capacity, connections, num_products, num_factories
        )
        
        # Calculate metrics
        total_demand = sum(demands)
        total_capacity = factory_capacity * num_factories
        fill_rate = (total_shipped / total_demand * 100) if total_demand > 0 else 0
        
        # Create row data
        row = {
            'Replication': rep + 1,
            'Seed': rep
        }
        
        # Add demand columns
        for i, demand in enumerate(demands):
            row[f'Demand_Product_{i+1}'] = round(demand, 2)
        
        # Add capacity columns
        for i in range(num_factories):
            row[f'Capacity_Factory_{i+1}'] = factory_capacity
        
        # Add flow columns
        flow_idx = 0
        for (product_idx, factory_idx), connected in connections.items():
            if connected:
                row[f'Flow_Product_{product_idx+1}_Factory_{factory_idx+1}'] = round(flows[flow_idx], 2)
                flow_idx += 1
        
        # Add summary columns
        row['Total_Demand'] = round(total_demand, 2)
        row['Total_Capacity'] = total_capacity
        row['Units_Sold'] = round(total_shipped, 2)
        row['Units_Lost'] = round(total_lost, 2)
        row['Fill_Rate'] = round(fill_rate, 2)
        
        results.append(row)
        
        # Update progress bar if provided
        if progress_bar is not None:
            progress_bar.progress((rep + 1) / num_replications)
    
    return pd.DataFrame(results)

def create_network_matrix(num_products: int, num_factories: int, connections: Dict[Tuple[int, int], bool]) -> pd.DataFrame:
    """Create a matrix representation of the network connections"""
    # Create matrix with products as rows and factories as columns
    matrix_data = []
    for product_idx in range(num_products):
        row = []
        for factory_idx in range(num_factories):
            # 1 if connected, 0 if not connected
            row.append(1 if connections.get((product_idx, factory_idx), False) else 0)
        matrix_data.append(row)
    
    # Create DataFrame with proper labels
    columns = [f"Factory {i+1}" for i in range(num_factories)]
    index = [f"Product {i+1}" for i in range(num_products)]
    
    return pd.DataFrame(matrix_data, columns=columns, index=index)

def create_summary_stats(results_df: pd.DataFrame) -> pd.DataFrame:
    """Create summary statistics from simulation results"""
    summary_data = {
        'Metric': [
            'Average Fill Rate (%)',
            'Std Dev Fill Rate (%)',
            'Min Fill Rate (%)',
            'Max Fill Rate (%)',
            'Average Units Sold',
            'Std Dev Units Sold',
            'Min Units Sold',
            'Max Units Sold',
            'Average Units Lost',
            'Std Dev Units Lost',
            'Min Units Lost',
            'Max Units Lost',
            'Total Replications'
        ],
        'Value': [
            round(results_df['Fill_Rate'].mean(), 2),
            round(results_df['Fill_Rate'].std(), 2),
            round(results_df['Fill_Rate'].min(), 2),
            round(results_df['Fill_Rate'].max(), 2),
            round(results_df['Units_Sold'].mean(), 2),
            round(results_df['Units_Sold'].std(), 2),
            round(results_df['Units_Sold'].min(), 2),
            round(results_df['Units_Sold'].max(), 2),
            round(results_df['Units_Lost'].mean(), 2),
            round(results_df['Units_Lost'].std(), 2),
            round(results_df['Units_Lost'].min(), 2),
            round(results_df['Units_Lost'].max(), 2),
            len(results_df)
        ]
    }
    return pd.DataFrame(summary_data)

def create_excel_file(results_df: pd.DataFrame, num_products: int, num_factories: int, 
                     connections: Dict[Tuple[int, int], bool], num_replications: int, progress_bar=None) -> bytes:
    """Create an Excel file with multiple sheets containing network layout, simulation results, and summary"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Update progress
    if progress_bar is not None:
        progress_bar.progress(0.1)
    
    # Create styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Network Layout
    ws1 = wb.create_sheet("Network Layout")
    ws1.title = "Network Layout"
    
    # Add title
    ws1['A1'] = "Network Configuration Matrix"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Products: {num_products}, Factorys: {num_factories}"
    ws1['A2'].font = Font(italic=True)
    
    # Create network matrix
    network_matrix = create_network_matrix(num_products, num_factories, connections)
    
    # Add matrix to sheet starting at row 4
    start_row = 4
    for r_idx, (product, row_data) in enumerate(network_matrix.iterrows()):
        ws1.cell(row=start_row + r_idx, column=1, value=product)
        for c_idx, value in enumerate(row_data):
            cell = ws1.cell(row=start_row + r_idx, column=c_idx + 2, value=value)
            if value == 1:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            cell.border = border
            cell.alignment = center_alignment
    
    # Add column headers
    for c_idx, col_name in enumerate(network_matrix.columns):
        cell = ws1.cell(row=start_row - 1, column=c_idx + 2, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Add row header
    ws1.cell(row=start_row - 1, column=1, value="Products")
    ws1.cell(row=start_row - 1, column=1).font = header_font
    ws1.cell(row=start_row - 1, column=1).fill = header_fill
    ws1.cell(row=start_row - 1, column=1).border = border
    ws1.cell(row=start_row - 1, column=1).alignment = center_alignment
    
    # Add legend
    legend_row = start_row + num_products + 2
    ws1.cell(row=legend_row, column=1, value="Legend:")
    ws1.cell(row=legend_row, column=1).font = Font(bold=True)
    ws1.cell(row=legend_row + 1, column=1, value="1 = Connected")
    ws1.cell(row=legend_row + 1, column=1).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    ws1.cell(row=legend_row + 2, column=1, value="0 = Not Connected")
    
    # Update progress after Sheet 1
    if progress_bar is not None:
        progress_bar.progress(0.4)
    
    # Sheet 2: Simulation Results
    ws2 = wb.create_sheet("Simulation Results")
    ws2.title = "Simulation Results"
    
    # Add title
    ws2['A1'] = f"Simulation Results ({num_replications} Replications)"
    ws2['A1'].font = Font(bold=True, size=14)
    
    # Add simulation data
    for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx + 2, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            cell.border = border
    
    # Update progress after Sheet 2
    if progress_bar is not None:
        progress_bar.progress(0.7)
    
    # Sheet 3: Summary Statistics
    ws3 = wb.create_sheet("Summary Statistics")
    ws3.title = "Summary Statistics"
    
    # Add title
    ws3['A1'] = "Summary Statistics"
    ws3['A1'].font = Font(bold=True, size=14)
    
    # Create summary data
    summary_df = create_summary_stats(results_df)
    
    # Add summary data
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx + 2, column=c_idx, value=value)
            if r_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            cell.border = border
    
    # Auto-adjust column widths
    for ws in [ws1, ws2, ws3]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Update progress before saving
    if progress_bar is not None:
        progress_bar.progress(0.9)
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Final progress update
    if progress_bar is not None:
        progress_bar.progress(1.0)
    
    return excel_buffer.getvalue()

# Main app
st.title("🏭 Process Flexibility Simulator")
st.markdown("Based on Jordan and Graves (1995): 'Principles on the Benefits of Manufacturing Process Flexibility'")

# Sidebar for inputs
# Use Streamlit's native info box for blue background
st.sidebar.info("**🏗️ Network Configuration**")

# Input fields
num_factories_products = st.sidebar.number_input("Number of Factorys and Products", min_value=1, max_value=10, value=6)
num_factories = num_factories_products
num_products = num_factories_products
factory_capacity = st.sidebar.number_input("Factory Capacity", min_value=1, max_value=1000, value=100)

# Demand distribution configuration
st.sidebar.subheader("Demand Distribution")
st.sidebar.markdown("**Truncated Normal Distribution Parameters:**")

# Demand distribution parameters
demand_mean = st.sidebar.number_input("Mean", min_value=1.0, max_value=1000.0, value=100.0, step=1.0)
demand_std = st.sidebar.number_input("Standard Deviation", min_value=1.0, max_value=500.0, value=40.0, step=1.0)
demand_min = st.sidebar.number_input("Lower Bound", min_value=1.0, max_value=1000.0, value=20.0, step=1.0)
demand_max = st.sidebar.number_input("Upper Bound", min_value=1.0, max_value=1000.0, value=180.0, step=1.0)
distribution_type = "Truncated Normal"

# Initialize connections if not exists
session_key = f'connections_{num_factories}_{num_products}'
if session_key not in st.session_state:
    st.session_state[session_key] = {}
    # Set "No Flexibility" as default only for new configurations
    default_connections = {}
    for product_idx in range(num_products):
        # Use modulo to cycle through factories, ensuring different assignments when possible
        factory_idx = product_idx % num_factories
        default_connections[(product_idx, factory_idx)] = True
    st.session_state[session_key] = default_connections

# Get connections with fallback
connections = st.session_state.get(session_key, {})

# Add connection toggle functionality
def toggle_connection(product_idx: int, factory_idx: int):
    """Toggle connection between product and factory"""
    key = (product_idx, factory_idx)
    current_connections = st.session_state.get(session_key, {})
    if key in current_connections:
        current_connections[key] = not current_connections[key]
    else:
        current_connections[key] = True
    st.session_state[session_key] = current_connections
    st.rerun()

# Add connection grid controls
st.sidebar.subheader("Connection Grid")
st.sidebar.write("Click buttons to toggle connections:")

# Add legend
st.sidebar.write("💡 **Legend**: 🔵 = Connected, ⚪ = Not Connected")

# Create a more compact grid view
if st.sidebar.checkbox("Compact Grid View", value=True):
    # Compact grid: Products as rows, Factorys as columns
    st.sidebar.write("**Products → Factorys**")
    
    # Header row with factory labels
    header_cols = st.sidebar.columns(num_factories + 1)
    with header_cols[0]:
        st.markdown("<small><strong>&nbsp;</strong></small>", unsafe_allow_html=True)
    for factory_idx in range(num_factories):
        with header_cols[factory_idx + 1]:
            st.markdown(f"<small><strong>F{factory_idx+1}</strong></small>", unsafe_allow_html=True)
    
    # Data rows
    for product_idx in range(num_products):
        row_cols = st.sidebar.columns(num_factories + 1)
        
        # Product label
        with row_cols[0]:
            st.markdown(f"<small><strong>P{product_idx+1}</strong></small>", unsafe_allow_html=True)
        
        # Buttons for each factory connection
        for factory_idx in range(num_factories):
            key = (product_idx, factory_idx)
            is_connected = connections.get(key, False)
            
            with row_cols[factory_idx + 1]:
                button_key = f"btn_{num_products}_{num_factories}_{product_idx}_{factory_idx}"
                button_text = "🔵" if is_connected else "⚪"
                if st.button(button_text, key=button_key, help=f"Toggle Product {product_idx+1} ↔ Factory {factory_idx+1}"):
                    current_connections = st.session_state.get(session_key, {})
                    current_connections[key] = not current_connections.get(key, False)
                    st.session_state[session_key] = current_connections
                    st.rerun()

else:
    # Detailed view: Each product gets its own section
    for product_idx in range(num_products):
        st.sidebar.write(f"**Product {product_idx+1}**")
        
        # Create columns for factories
        cols = st.sidebar.columns(num_factories)
        
        for factory_idx, col in enumerate(cols):
            key = (product_idx, factory_idx)
            is_connected = connections.get(key, False)
            
            with col:
                button_text = f"F{factory_idx+1} 🔵" if is_connected else f"F{factory_idx+1} ⚪"
                button_key = f"detailed_{num_products}_{num_factories}_{product_idx}_{factory_idx}"
                if st.button(button_text, key=button_key, help=f"Toggle Product {product_idx+1} ↔ Factory {factory_idx+1}"):
                    current_connections = st.session_state.get(session_key, {})
                    current_connections[key] = not current_connections.get(key, False)
                    st.session_state[session_key] = current_connections
                    st.rerun()
        
        st.sidebar.write("---")  # Separator between products

# Quick setup buttons
st.sidebar.subheader("Quick Setup")

# First row: Main flexibility options
col_a, col_b, col_c = st.sidebar.columns(3)

with col_a:
    if st.button("Full Flexibility", help="Connect all products to all factories", key=f"btn_full_{num_products}_{num_factories}"):
        # Create new connections dictionary
        new_connections = {}
        for product_idx in range(num_products):
            for factory_idx in range(num_factories):
                new_connections[(product_idx, factory_idx)] = True
        st.session_state[session_key] = new_connections
        # Clear previous simulation results since network configuration changed
        if 'simulation_completed' in st.session_state:
            del st.session_state.simulation_completed
        if 'simulation_results' in st.session_state:
            del st.session_state.simulation_results
        if 'excel_ready' in st.session_state:
            del st.session_state.excel_ready
        if 'excel_data' in st.session_state:
            del st.session_state.excel_data
        if 'excel_filename' in st.session_state:
            del st.session_state.excel_filename
        st.rerun()

with col_b:
    if st.button("No Flexibility", help="Connect each product to one factory (dedicated assignment)", key=f"btn_none_{num_products}_{num_factories}"):
        # Create new connections dictionary
        new_connections = {}
        # Connect each product to a single factory, preferably different ones
        for product_idx in range(num_products):
            # Use modulo to cycle through factories, ensuring different assignments when possible
            factory_idx = product_idx % num_factories
            new_connections[(product_idx, factory_idx)] = True
        st.session_state[session_key] = new_connections
        # Clear previous simulation results since network configuration changed
        if 'simulation_completed' in st.session_state:
            del st.session_state.simulation_completed
        if 'simulation_results' in st.session_state:
            del st.session_state.simulation_results
        if 'excel_ready' in st.session_state:
            del st.session_state.excel_ready
        if 'excel_data' in st.session_state:
            del st.session_state.excel_data
        if 'excel_filename' in st.session_state:
            del st.session_state.excel_filename
        st.rerun()
    

with col_c:
    if st.button("Remove Connections", help="Remove all connections", key=f"btn_remove_{num_products}_{num_factories}"):
        # Create empty connections dictionary
        st.session_state[session_key] = {}
        # Clear previous simulation results since network configuration changed
        if 'simulation_completed' in st.session_state:
            del st.session_state.simulation_completed
        if 'simulation_results' in st.session_state:
            del st.session_state.simulation_results
        if 'excel_ready' in st.session_state:
            del st.session_state.excel_ready
        if 'excel_data' in st.session_state:
            del st.session_state.excel_data
        if 'excel_filename' in st.session_state:
            del st.session_state.excel_filename
        st.rerun()

# Second row: 2-Flexibility and 3-Flexibility buttons (if enabled)
if SHOW_ADVANCED_FLEXIBILITY_OPTIONS:
    col_d, col_e, col_f = st.sidebar.columns(3)
    
    with col_d:
        if st.button("2-Flexibility", help="Create 2-chain flexibility pattern", key=f"btn_2flex_{num_products}_{num_factories}"):
            # Create new connections dictionary
            new_connections = {}
            # Create 2-chain pattern: P1→Factory1,Factory2; P2→Factory2,Factory3; P3→Factory3,Factory1; etc.
            for product_idx in range(num_products):
                # Each product connects to 2 factories in a chain pattern
                factory1_idx = product_idx % num_factories
                factory2_idx = (product_idx + 1) % num_factories
                new_connections[(product_idx, factory1_idx)] = True
                new_connections[(product_idx, factory2_idx)] = True
            st.session_state[session_key] = new_connections
            # Clear previous simulation results since network configuration changed
            if 'simulation_completed' in st.session_state:
                del st.session_state.simulation_completed
            if 'simulation_results' in st.session_state:
                del st.session_state.simulation_results
            if 'excel_ready' in st.session_state:
                del st.session_state.excel_ready
            if 'excel_data' in st.session_state:
                del st.session_state.excel_data
            if 'excel_filename' in st.session_state:
                del st.session_state.excel_filename
            st.rerun()

    with col_e:
        if st.button("3-Flexibility", help="Create 3-chain flexibility pattern", key=f"btn_3flex_{num_products}_{num_factories}"):
            # Create new connections dictionary
            new_connections = {}
            # Create 3-chain pattern: P1→Factory1,Factory2,Factory3; P2→Factory2,Factory3,Factory1; etc.
            for product_idx in range(num_products):
                # Each product connects to 3 factories in a chain pattern
                factory1_idx = product_idx % num_factories
                factory2_idx = (product_idx + 1) % num_factories
                factory3_idx = (product_idx + 2) % num_factories
                new_connections[(product_idx, factory1_idx)] = True
                new_connections[(product_idx, factory2_idx)] = True
                new_connections[(product_idx, factory3_idx)] = True
            st.session_state[session_key] = new_connections
            # Clear previous simulation results since network configuration changed
            if 'simulation_completed' in st.session_state:
                del st.session_state.simulation_completed
            if 'simulation_results' in st.session_state:
                del st.session_state.simulation_results
            if 'excel_ready' in st.session_state:
                del st.session_state.excel_ready
            if 'excel_data' in st.session_state:
                del st.session_state.excel_data
            if 'excel_filename' in st.session_state:
                del st.session_state.excel_filename
            st.rerun()





# Main content area - just the graph
# Refresh connections from session state to ensure consistency
connections = st.session_state.get(session_key, {})

# Create the graph
fig = create_bipartite_graph(num_factories, num_products, connections)
st.plotly_chart(fig, use_container_width=True)

# Instructions
st.info("💡 **Instructions**: Use the checkboxes in the sidebar to toggle connections between products and factories. Solid blue lines indicate active connections (flexibility).")

# Rerun handling is now done directly in button handlers


# Add horizontal divider
st.sidebar.markdown("---")

# Use Streamlit's native success box for green background
st.sidebar.success("**📊 Analysis & Simulation**")

# Replication input and simulator button
st.sidebar.subheader("Simulation Parameters")
num_replications = st.sidebar.number_input("Number of Replications", min_value=1, max_value=10000, value=100, step=100)

# Add specific CSS for Run Simulator button
st.markdown("""
<style>
    div[data-testid="stSidebar"] button[data-testid="baseButton-primary"] {
        width: 100% !important;
        min-width: 200px !important;
        max-width: none !important;
        font-size: 1rem !important;
        padding: 0.75rem 1.5rem !important;
    }
    
    /* Alternative approach - target by button text content */
    button:has-text("Run Simulator") {
        width: 100% !important;
        min-width: 200px !important;
    }
    
    /* Even more specific targeting */
    .stButton > button[data-testid="baseButton-primary"] {
        width: 100% !important;
        min-width: 200px !important;
        max-width: none !important;
    }
</style>
""", unsafe_allow_html=True)

if st.sidebar.button("🚀 Run Simulator", type="primary", key="run_simulator_btn"):
    # Create progress bar for simulation
    progress_bar = st.sidebar.progress(0)
    status_text = st.sidebar.empty()
    
    # Run the simulation
    demand_params = {
        'mean': demand_mean,
        'std': demand_std,
        'min': demand_min,
        'max': demand_max
    }
    
    status_text.text("Running simulation...")
    simulation_results = run_simulation(
        num_replications, num_products, num_factories, factory_capacity, connections, demand_params, progress_bar
    )
    
    # Store results in session state
    st.session_state.simulation_results = simulation_results
    st.session_state.simulation_completed = True
    
    # Clear progress bar and show success
    progress_bar.empty()
    status_text.empty()
    st.sidebar.success(f"Simulation completed with {num_replications} replications!")

# Display simulation results and download button
if st.session_state.get('simulation_completed', False):
    st.subheader("Simulation Results")
    
    # Display summary statistics
    results_df = st.session_state.simulation_results
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Average Fill Rate", f"{results_df['Fill_Rate'].mean():.1f}%")
    with col2:
        avg_sold = results_df['Units_Sold'].mean()
        std_sold = results_df['Units_Sold'].std()
        st.metric("Average Units Sold", f"{avg_sold:.1f} ± {std_sold:.1f}")
    with col3:
        avg_lost = results_df['Units_Lost'].mean()
        std_lost = results_df['Units_Lost'].std()
        st.metric("Average Units Lost", f"{avg_lost:.1f} ± {std_lost:.1f}")
    with col4:
        st.metric("Total Replications", len(results_df))
    
    # Display first few rows
    st.subheader("Sample Results (First 10 Replications)")
    st.dataframe(results_df.head(10), use_container_width=True)
    
    # Excel download section
    if st.button("📊 Generate Excel File", type="primary", help="Generate Excel file with network layout, simulation results, and summary statistics"):
        # Create progress bar for Excel generation
        excel_progress = st.progress(0)
        excel_status = st.empty()
        
        # Count number of selected edges (connections)
        num_edges = sum(1 for connected in connections.values() if connected)
        
        # Generate timestamp
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        excel_status.text("Generating Excel file...")
        excel_data = create_excel_file(results_df, num_products, num_factories, connections, num_replications, excel_progress)
        
        # Clear progress indicators
        excel_progress.empty()
        excel_status.empty()
        
        # Store Excel data in session state
        st.session_state.excel_data = excel_data
        st.session_state.excel_filename = f"{timestamp}_flexibility_simulation_{num_products}products_{num_factories}factories_{num_edges}edges_{num_replications}reps.xlsx"
        st.session_state.excel_ready = True
        
        st.success("Excel file generated successfully!")
    
    # Show download button if Excel file is ready
    if st.session_state.get('excel_ready', False):
        st.download_button(
            label="📥 Download Excel Results",
            data=st.session_state.excel_data,
            file_name=st.session_state.excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            help="Download the generated Excel file"
        )

# Display demand data if available
if st.session_state.demand_data:
    st.subheader("Demand Distribution Visualization")
    
    # Create demand plots
    fig_demand = go.Figure()
    
    for product_idx, demands in st.session_state.demand_data.items():
        fig_demand.add_trace(go.Histogram(
            x=demands,
            name=f'Product {product_idx}',
            opacity=0.7
        ))
    
    fig_demand.update_layout(
        title="Demand Distribution by Product",
        xaxis_title="Demand",
        yaxis_title="Frequency",
        barmode='overlay'
    )
    
    st.plotly_chart(fig_demand, use_container_width=True)
    
    # Display summary statistics
    st.subheader("Demand Summary Statistics")
    summary_data = []
    for product_idx, demands in st.session_state.demand_data.items():
        summary_data.append({
            'Product': f'Product {product_idx}',
            'Mean': np.mean(demands),
            'Std': np.std(demands),
            'Min': np.min(demands),
            'Max': np.max(demands)
        })
    
    df_summary = pd.DataFrame(summary_data)
    st.dataframe(df_summary, use_container_width=True)
    
    # Optimization results
    st.subheader("Optimization Results")
    
    # Calculate theoretical benefits
    total_demand = sum([np.mean(demands) for demands in st.session_state.demand_data.values()])
    total_capacity = factory_capacity * num_factories
    
    col_opt1, col_opt2, col_opt3 = st.columns(3)
    
    with col_opt1:
        st.metric("Total Demand", f"{total_demand:.0f}")
    
    with col_opt2:
        st.metric("Total Capacity", f"{total_capacity:.0f}")
    
    with col_opt3:
        utilization = (total_demand / total_capacity) * 100 if total_capacity > 0 else 0
        st.metric("System Utilization", f"{utilization:.1f}%")
    
    # Flexibility benefits analysis
    st.subheader("Flexibility Benefits Analysis")
    
    # Calculate flexibility metrics
    metrics = calculate_flexibility_metrics(connections, num_products, num_factories)
    
    # Calculate flexibility benefits based on Jordan & Graves principles
    flexibility_benefits = {
        'Risk Pooling': f"{(metrics['flexibility_ratio'] * 100):.1f}% risk reduction",
        'Demand Smoothing': f"{(1 - metrics['flexibility_ratio']) * 50:.1f}% demand variability reduction",
        'Capacity Utilization': f"{min(100, utilization + (metrics['flexibility_ratio'] * 20)):.1f}%"
    }
    
    for benefit, value in flexibility_benefits.items():
        st.metric(benefit, value)

# Credits section at the bottom of sidebar
st.sidebar.markdown("---")
st.sidebar.warning("**📝 Credits**")

st.sidebar.markdown("**Created by:** Yaron Shaposhnik")
st.sidebar.markdown("**Email:** yaron.shaposhnik@simon.rochester.edu")
st.sidebar.markdown("**Year:** 2025")
st.sidebar.markdown("*Process Flexibility Simulation Tool*")


